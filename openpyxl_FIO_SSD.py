# -*- coding: utf-8 -*-
"""
该脚本文件是对fio的SSD日志文件内的时间与数据的之和进行处理操作并生成excel文件汇总
"""
import os
import openpyxl
# import major_function
file_path = '/results_QPS_data'     # 日志文件存放的路径
# save_path = './'   # 文件保存路径当前文件夹下
# excel_name = 'fio_ssd_bw_sun(11_01)'     # 保存的excel文件名称

path = os.path.split(os.path.realpath(__file__))[0]
print(path)
if os.path.isdir(path + '/save_csv'):
    file_list = os.listdir(path + '/save_csv')
    if file_list:
        for file_name in file_list:
            os.remove(path + '/save_csv' + '/' + file_name)
else:
    os.mkdir(path + '/save_csv')

save_path = path + '/save_csv'  # csv文件保存路径当前文件夹下


def sequence_fiossd():
    folder_names = os.listdir(path + file_path)
    print(folder_names)
    for folder in folder_names:
        file_names = os.listdir(path + file_path + '/' + folder)
        print(file_names)
        # workbook = major_function.Instantiate_csv_object()
        workbook = openpyxl.Workbook()
        ws = workbook.active
        #更改工作表ws的title
        ws.title = 'test_sheet1'
        # sheet = ws.title(index=0, title='sheet1')
        # sheet.write(0, 0, 'times')
        time_list = []
        with open(path + file_path + '/' + folder + '/' + file_names[0], 'r') as f:
            for line in f.readlines():
                time_list.append(line.split(',')[0].split(' ')[0])
            f.close
            print(time_list)
            for s in range(len(time_list)):
                ws.cell(s+2, 1, int(time_list[s]))
        data_list_all = []
        for i in range(len(file_names)):
            data_list = []
            ws.cell(1, 1, 'time/ms')
            ws.cell(1, i+2, 'log' + str(i+1))
            with open(path + file_path + '/' + folder + '/' + file_names[i], 'r') as f:
                for line in f.readlines():
                    # print(len(line))
                    if len(line) != 1:
                        # print(line.split(',')[1].split(' '))
                        data_list.append(line.split(',')[1])
                f.close
                data_list_all.append(data_list)
                for a in range(len(data_list)):
                    data_list[a]
                    ws.cell(a+2, i+2, int(data_list[a]))

        sum_list = []
        sum_list2 = []
        file_list = []
        for a in range(len(data_list_all)):
            file_list.append(len(data_list_all[a]))

        try:
            for l in range(min(file_list)):
                sum = 0
                for j in range(len(data_list_all)):
                    sum = (int(data_list_all[j][l]) + sum)
                sum_list2.append(sum)
                sum_list.append(sum/1000)
        except Exception as err:
            print(err)
            break

        col = len(data_list_all)
        ws.cell(1, col+2, 'sum')
        for ks in range(len(sum_list2)):
            ws.cell(ks+2, col+2, int(sum_list2[ks]))
        ws.cell(1, col+3, 'sum/1000')
        for k in range(len(sum_list)):
            ws.cell(k+2, col+3, int(sum_list[k]))
        workbook.save(save_path + '/' + folder + '.csv')


if __name__ == '__main__':
    # pass
    sequence_fiossd()
